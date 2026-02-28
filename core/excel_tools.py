"""
Excel Tool Registry — Outils pour créer, lire et modifier des fichiers Excel (.xlsx).

Utilise openpyxl (pure Python, compatible Windows/Linux).
Maintient un cache en mémoire des workbooks ouverts pour éviter les lectures disque répétées.

Outils disponibles :
  create_excel      — Créer un nouveau classeur
  open_excel        — Ouvrir un classeur existant
  read_sheet        — Lire les données d'une feuille
  write_cell        — Écrire une valeur dans une cellule
  write_rows        — Écrire plusieurs lignes
  add_sheet         — Ajouter une feuille
  delete_sheet      — Supprimer une feuille
  list_sheets       — Lister les feuilles
  format_cells      — Formater une plage de cellules
  apply_formula     — Insérer une formule Excel
  auto_fit_columns  — Ajuster automatiquement la largeur des colonnes
  save_excel        — Sauvegarder le classeur
  get_excel_info    — Obtenir les métadonnées du classeur
"""
from pathlib import Path
from typing import Any, Dict, List, Optional, TYPE_CHECKING

try:
    import openpyxl
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

if TYPE_CHECKING:
    from core.memory import MemoryManager


# --------------------------------------------------------------------------- #
#  Tool definitions                                                            #
# --------------------------------------------------------------------------- #

EXCEL_TOOL_DEFINITIONS: List[Dict] = [
    {
        "name": "create_excel",
        "description": "Create a new Excel workbook (.xlsx) at the given path. Returns the file path.",
        "params": {
            "path": "Path to the Excel file to create (e.g. 'C:/data/report.xlsx')",
            "sheet_name": "Name of the initial sheet (default: 'Sheet1')",
        },
        "required": ["path"],
    },
    {
        "name": "open_excel",
        "description": "Open an existing Excel workbook and return its structure (sheets, dimensions).",
        "params": {
            "path": "Path to the Excel file",
        },
        "required": ["path"],
    },
    {
        "name": "read_sheet",
        "description": "Read data from a sheet in an Excel file. Returns rows as a list of dicts (or lists if no header).",
        "params": {
            "path": "Path to the Excel file",
            "sheet_name": "Sheet name (default: active/first sheet)",
            "has_header": "First row is a header row (default: true)",
            "min_row": "Start row number 1-based (optional)",
            "max_row": "End row number 1-based (optional)",
        },
        "required": ["path"],
    },
    {
        "name": "write_cell",
        "description": "Write a value or formula to a specific cell in a sheet.",
        "params": {
            "path": "Path to the Excel file",
            "cell": "Cell reference (e.g. 'A1', 'B3')",
            "value": "Value to write — string, number, or Excel formula starting with '='",
            "sheet_name": "Sheet name (default: active sheet)",
        },
        "required": ["path", "cell", "value"],
    },
    {
        "name": "write_rows",
        "description": "Write multiple rows of data to a sheet. 'rows' is a list of lists.",
        "params": {
            "path": "Path to the Excel file",
            "rows": "List of rows, each row is a list of values",
            "sheet_name": "Sheet name (default: active sheet)",
            "start_row": "Starting row number 1-based (default: 1)",
            "start_col": "Starting column number 1-based (default: 1)",
        },
        "required": ["path", "rows"],
    },
    {
        "name": "add_sheet",
        "description": "Add a new sheet to an existing Excel workbook.",
        "params": {
            "path": "Path to the Excel file",
            "sheet_name": "Name for the new sheet",
            "position": "Insert position — 0-based index (optional, appended by default)",
        },
        "required": ["path", "sheet_name"],
    },
    {
        "name": "delete_sheet",
        "description": "Delete a sheet from an Excel workbook.",
        "params": {
            "path": "Path to the Excel file",
            "sheet_name": "Name of the sheet to delete",
        },
        "required": ["path", "sheet_name"],
    },
    {
        "name": "list_sheets",
        "description": "List all sheet names in an Excel workbook.",
        "params": {
            "path": "Path to the Excel file",
        },
        "required": ["path"],
    },
    {
        "name": "format_cells",
        "description": (
            "Apply formatting to a range of cells: bold, font color, background color, "
            "text alignment, font size."
        ),
        "params": {
            "path": "Path to the Excel file",
            "range": "Cell range (e.g. 'A1:E1', 'A1', 'B2:D5')",
            "sheet_name": "Sheet name (default: active sheet)",
            "bold": "Make text bold — true or false",
            "font_color": "Font color as ARGB hex (e.g. 'FF0000' for red)",
            "bg_color": "Background fill color as ARGB hex (e.g. 'FFFF00' for yellow)",
            "alignment": "Text alignment: 'left', 'center', or 'right'",
            "font_size": "Font size in points (e.g. 12)",
        },
        "required": ["path", "range"],
    },
    {
        "name": "apply_formula",
        "description": "Insert an Excel formula into a cell (e.g. '=SUM(A1:B1)').",
        "params": {
            "path": "Path to the Excel file",
            "cell": "Target cell reference (e.g. 'C10')",
            "formula": "Excel formula — with or without leading '=' (it will be added automatically)",
            "sheet_name": "Sheet name (default: active sheet)",
        },
        "required": ["path", "cell", "formula"],
    },
    {
        "name": "auto_fit_columns",
        "description": "Auto-fit column widths to match the widest content in each column.",
        "params": {
            "path": "Path to the Excel file",
            "sheet_name": "Sheet name (default: active sheet)",
        },
        "required": ["path"],
    },
    {
        "name": "save_excel",
        "description": "Save the workbook to disk. Optionally save to a different path (Save As).",
        "params": {
            "path": "Original workbook path",
            "save_as": "Alternative save path for Save As (optional)",
        },
        "required": ["path"],
    },
    {
        "name": "get_excel_info",
        "description": "Get metadata about an Excel file: sheet names, row/column counts, dimensions.",
        "params": {
            "path": "Path to the Excel file",
        },
        "required": ["path"],
    },
    {
        "name": "store_finding",
        "description": "Store an important finding in semantic memory for later retrieval.",
        "params": {
            "key": "Short identifier for this finding",
            "value": "The finding content (text or structured data)",
            "category": "Category: 'excel', 'data', 'finding'",
        },
        "required": ["key", "value"],
    },
    {
        "name": "think",
        "description": "Pure reasoning step — no external action. Use to plan before acting.",
        "params": {"reasoning": "Your detailed reasoning"},
        "required": ["reasoning"],
    },
    {
        "name": "final_answer",
        "description": "Emit the final answer when the task is complete.",
        "params": {
            "answer": "The complete answer or report",
            "summary": "One-sentence executive summary",
        },
        "required": ["answer"],
    },
]

EXCEL_TOOL_NAMES = {t["name"] for t in EXCEL_TOOL_DEFINITIONS}


# --------------------------------------------------------------------------- #
#  Tool executor                                                               #
# --------------------------------------------------------------------------- #

class ExcelToolExecutor:
    """
    Executes Excel tools using openpyxl.
    Maintains an in-memory workbook cache to avoid repeated disk reads/writes.
    """

    def __init__(
        self,
        memory: "MemoryManager",
        dispatch_callback=None,
    ):
        if not OPENPYXL_AVAILABLE:
            raise ImportError(
                "openpyxl is required for ExcelAgent. "
                "Install it with: pip install openpyxl"
            )
        self.memory = memory
        self._dispatch = dispatch_callback
        self._wb_cache: Dict[str, Any] = {}   # abs_path → Workbook

    def execute(self, tool_name: str, params: Dict[str, Any]) -> Any:
        if tool_name not in EXCEL_TOOL_NAMES:
            raise ValueError(f"Unknown Excel tool: {tool_name}")
        method = getattr(self, f"_tool_{tool_name}", None)
        if method is None:
            raise ValueError(f"Tool '{tool_name}' is defined but not implemented")
        return method(**params)

    # ------------------------------------------------------------------ #
    #  Internal helpers                                                    #
    # ------------------------------------------------------------------ #

    def _load_wb(self, path: str):
        """Load workbook from cache or disk; create new if file does not exist."""
        abs_path = str(Path(path).resolve())
        if abs_path not in self._wb_cache:
            if Path(abs_path).exists():
                self._wb_cache[abs_path] = load_workbook(abs_path)
            else:
                wb = Workbook()
                self._wb_cache[abs_path] = wb
        return self._wb_cache[abs_path], abs_path

    def _save_wb(self, wb: Any, path: str) -> None:
        wb.save(path)

    def _get_sheet(self, wb: Any, sheet_name: Optional[str] = None):
        if sheet_name and sheet_name in wb.sheetnames:
            return wb[sheet_name]
        return wb.active

    # ------------------------------------------------------------------ #
    #  Excel tools                                                         #
    # ------------------------------------------------------------------ #

    def _tool_create_excel(self, path: str, sheet_name: str = "Sheet1") -> str:
        abs_path = str(Path(path).resolve())
        Path(abs_path).parent.mkdir(parents=True, exist_ok=True)
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        wb.save(abs_path)
        self._wb_cache[abs_path] = wb
        return f"Excel workbook created: {abs_path}"

    def _tool_open_excel(self, path: str) -> Dict:
        wb, abs_path = self._load_wb(path)
        sheets_info = {}
        for sname in wb.sheetnames:
            ws = wb[sname]
            sheets_info[sname] = {
                "dimensions": ws.dimensions,
                "max_row": ws.max_row,
                "max_column": ws.max_column,
            }
        return {"path": abs_path, "sheet_count": len(wb.sheetnames), "sheets": sheets_info}

    def _tool_read_sheet(
        self,
        path: str,
        sheet_name: Optional[str] = None,
        has_header: bool = True,
        min_row: Optional[int] = None,
        max_row: Optional[int] = None,
    ) -> List:
        wb, _ = self._load_wb(path)
        ws = self._get_sheet(wb, sheet_name)
        rows = list(ws.iter_rows(min_row=min_row, max_row=max_row, values_only=True))
        if not rows:
            return []
        if has_header:
            headers = [
                str(h) if h is not None else f"col_{i}"
                for i, h in enumerate(rows[0])
            ]
            return [dict(zip(headers, row)) for row in rows[1:]]
        return [list(row) for row in rows]

    def _tool_write_cell(
        self,
        path: str,
        cell: str,
        value: Any,
        sheet_name: Optional[str] = None,
    ) -> str:
        wb, abs_path = self._load_wb(path)
        ws = self._get_sheet(wb, sheet_name)
        ws[cell] = value
        self._save_wb(wb, abs_path)
        return f"Cell {cell} set to '{value}' in sheet '{ws.title}' ({abs_path})"

    def _tool_write_rows(
        self,
        path: str,
        rows: List[List],
        sheet_name: Optional[str] = None,
        start_row: int = 1,
        start_col: int = 1,
    ) -> str:
        wb, abs_path = self._load_wb(path)
        ws = self._get_sheet(wb, sheet_name)
        for r_idx, row_data in enumerate(rows):
            for c_idx, val in enumerate(row_data):
                ws.cell(row=start_row + r_idx, column=start_col + c_idx, value=val)
        self._save_wb(wb, abs_path)
        return f"{len(rows)} rows written to sheet '{ws.title}' in {abs_path}"

    def _tool_add_sheet(
        self,
        path: str,
        sheet_name: str,
        position: Optional[int] = None,
    ) -> str:
        wb, abs_path = self._load_wb(path)
        if sheet_name in wb.sheetnames:
            return f"Sheet '{sheet_name}' already exists in {abs_path}."
        wb.create_sheet(title=sheet_name, index=position)
        self._save_wb(wb, abs_path)
        return f"Sheet '{sheet_name}' added to {abs_path}"

    def _tool_delete_sheet(self, path: str, sheet_name: str) -> str:
        wb, abs_path = self._load_wb(path)
        if sheet_name not in wb.sheetnames:
            return f"Sheet '{sheet_name}' not found in {abs_path}."
        del wb[sheet_name]
        self._save_wb(wb, abs_path)
        return f"Sheet '{sheet_name}' deleted from {abs_path}"

    def _tool_list_sheets(self, path: str) -> List[str]:
        wb, _ = self._load_wb(path)
        return list(wb.sheetnames)

    def _tool_format_cells(
        self,
        path: str,
        range: str,
        sheet_name: Optional[str] = None,
        bold: bool = False,
        font_color: Optional[str] = None,
        bg_color: Optional[str] = None,
        alignment: Optional[str] = None,
        font_size: Optional[int] = None,
    ) -> str:
        wb, abs_path = self._load_wb(path)
        ws = self._get_sheet(wb, sheet_name)
        cell_range = ws[range]
        # Normalize to iterable of rows
        if not isinstance(cell_range, (list, tuple)):
            cell_range = [[cell_range]]
        for row in cell_range:
            if not isinstance(row, (list, tuple)):
                row = [row]
            for cell in row:
                if bold or font_color or font_size:
                    cell.font = Font(
                        bold=bold,
                        color=font_color,
                        size=font_size,
                    )
                if bg_color:
                    cell.fill = PatternFill(fill_type="solid", fgColor=bg_color)
                if alignment:
                    cell.alignment = Alignment(horizontal=alignment)
        self._save_wb(wb, abs_path)
        return f"Formatting applied to '{range}' in sheet '{ws.title}'"

    def _tool_apply_formula(
        self,
        path: str,
        cell: str,
        formula: str,
        sheet_name: Optional[str] = None,
    ) -> str:
        wb, abs_path = self._load_wb(path)
        ws = self._get_sheet(wb, sheet_name)
        ws[cell] = formula if formula.startswith("=") else f"={formula}"
        self._save_wb(wb, abs_path)
        return f"Formula '{formula}' written to cell {cell} in sheet '{ws.title}'"

    def _tool_auto_fit_columns(
        self,
        path: str,
        sheet_name: Optional[str] = None,
    ) -> str:
        wb, abs_path = self._load_wb(path)
        ws = self._get_sheet(wb, sheet_name)
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value is not None:
                        max_len = max(max_len, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 2, 60)
        self._save_wb(wb, abs_path)
        return f"Column widths auto-fitted in sheet '{ws.title}'"

    def _tool_save_excel(self, path: str, save_as: Optional[str] = None) -> str:
        wb, abs_path = self._load_wb(path)
        target = str(Path(save_as).resolve()) if save_as else abs_path
        if save_as:
            Path(target).parent.mkdir(parents=True, exist_ok=True)
        wb.save(target)
        if save_as:
            self._wb_cache[target] = wb
        return f"Workbook saved to {target}"

    def _tool_get_excel_info(self, path: str) -> Dict:
        wb, abs_path = self._load_wb(path)
        sheets_info = []
        for sname in wb.sheetnames:
            ws = wb[sname]
            sheets_info.append({
                "name": sname,
                "max_row": ws.max_row,
                "max_column": ws.max_column,
                "dimensions": ws.dimensions,
            })
        return {
            "path": abs_path,
            "sheet_count": len(wb.sheetnames),
            "sheets": sheets_info,
        }

    # ------------------------------------------------------------------ #
    #  Memory & system tools                                               #
    # ------------------------------------------------------------------ #

    def _tool_store_finding(
        self,
        key: str,
        value: Any,
        category: str = "finding",
    ) -> str:
        self.memory.store_fact(key, value, source="excel-agent",
                               category=category, confidence=1.0)
        return f"Finding '{key}' stored in semantic memory."

    def _tool_think(self, reasoning: str = "") -> str:
        return f"[REASONING] {reasoning}"

    def _tool_final_answer(self, answer: str, summary: str = "") -> Dict[str, str]:
        return {"answer": answer, "summary": summary or answer[:200]}
